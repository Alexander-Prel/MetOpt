from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from math import ceil
from typing import Dict, List, Tuple

from pathlib import Path
from openpyxl import load_workbook

ASSET_NAMES = {
    "цб 1": "cb1",
    "цб1": "cb1",
    "цб 2": "cb2",
    "цб2": "cb2",
    "депозиты": "dep",
    "депозит": "dep",
}

UNIT = 25  # базовая дискретизация денег
CASH_ROUND = 200  # округление свободных средств (д.е.) для уменьшения числа состояний


@dataclass(frozen=True)
class Scenario:
    prob: int   # percent (0..100)
    m_cb1: int  # multiplier * 100
    m_cb2: int
    m_dep: int


def normalize_name(raw: str) -> str:
    return " ".join(str(raw).strip().lower().split())


def to_int_percent(value) -> int:
    return int(round(float(value) * 100))


def find_row(ws, col: int, text: str) -> int:
    target = normalize_name(text)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        if normalize_name(v) == target:
            return r
    raise ValueError(f"Не найдена строка: {text}")


def read_excel(path: str) -> Dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Probabilities for scenarios (rows 4-6, columns K-M)
    probs = {
        "fav": [ws.cell(4, 11).value, ws.cell(4, 12).value, ws.cell(4, 13).value],
        "neu": [ws.cell(5, 11).value, ws.cell(5, 12).value, ws.cell(5, 13).value],
        "neg": [ws.cell(6, 11).value, ws.cell(6, 12).value, ws.cell(6, 13).value],
    }

    # Asset blocks (column I has asset name, column J initial value, K-M multipliers)
    asset_rows: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 9).value  # column I
        if val is None:
            continue
        name = normalize_name(val)
        if name in ASSET_NAMES:
            key = ASSET_NAMES[name]
            asset_rows[key] = r

    if not asset_rows:
        raise ValueError("Не удалось найти блоки ЦБ/депозитов в файле.")

    def read_asset_block(key: str) -> Tuple[float, List[List[int]]]:
        r = asset_rows[key]
        init = float(ws.cell(r, 10).value)
        mults = [
            [to_int_percent(ws.cell(r, 11).value), to_int_percent(ws.cell(r, 12).value), to_int_percent(ws.cell(r, 13).value)],
            [to_int_percent(ws.cell(r + 1, 11).value), to_int_percent(ws.cell(r + 1, 12).value), to_int_percent(ws.cell(r + 1, 13).value)],
            [to_int_percent(ws.cell(r + 2, 11).value), to_int_percent(ws.cell(r + 2, 12).value), to_int_percent(ws.cell(r + 2, 13).value)],
        ]
        return init, mults

    cb1_init, cb1_mults = read_asset_block("cb1")
    cb2_init, cb2_mults = read_asset_block("cb2")
    dep_init, dep_mults = read_asset_block("dep")

    # Commissions block
    row_comm = find_row(ws, 3, "Комиссии брокеров")
    commissions: Dict[str, int] = {}
    for r in range(row_comm + 1, row_comm + 5):
        name = ws.cell(r, 3).value
        rate = ws.cell(r, 4).value
        if name is None or rate is None:
            continue
        nm = normalize_name(name)
        if nm in ASSET_NAMES:
            commissions[ASSET_NAMES[nm]] = to_int_percent(rate)

    # Minimum holdings block
    row_min = find_row(ws, 3, "Не менее")
    minimums: Dict[str, float] = {}
    for r in range(row_min + 1, row_min + 5):
        name = ws.cell(r, 3).value
        val = ws.cell(r, 4).value
        if name is None or val is None:
            continue
        nm = normalize_name(name)
        if nm in ASSET_NAMES:
            minimums[ASSET_NAMES[nm]] = float(val)

    # Build scenarios per stage (3 stages)
    stages: List[List[Scenario]] = []
    for stage_idx in range(3):
        fav = Scenario(
            prob=to_int_percent(probs["fav"][stage_idx]),
            m_cb1=cb1_mults[0][stage_idx],
            m_cb2=cb2_mults[0][stage_idx],
            m_dep=dep_mults[0][stage_idx],
        )
        neu = Scenario(
            prob=to_int_percent(probs["neu"][stage_idx]),
            m_cb1=cb1_mults[1][stage_idx],
            m_cb2=cb2_mults[1][stage_idx],
            m_dep=dep_mults[1][stage_idx],
        )
        neg = Scenario(
            prob=to_int_percent(probs["neg"][stage_idx]),
            m_cb1=cb1_mults[2][stage_idx],
            m_cb2=cb2_mults[2][stage_idx],
            m_dep=dep_mults[2][stage_idx],
        )
        stages.append([fav, neu, neg])

    return {
        "stages": stages,
        "init": {"cb1": cb1_init, "cb2": cb2_init, "dep": dep_init},
        "comm": commissions,
        "mins": minimums,
    }


@dataclass(frozen=True)
class Action:
    k1: int
    k2: int
    k3: int


def apply_multiplier_to_packages(packages: int, pkg_value: float, m: int) -> int:
    value = packages * pkg_value
    value = (value * m + 50) // 100
    return int(round(value / pkg_value))


def round_cash_units(cash_units: int) -> int:
    step = CASH_ROUND // UNIT
    return int(round(cash_units / step) * step)


def action_to_text(action: Action) -> str:
    parts = []
    if action.k1:
        parts.append(f"ЦБ1: {'купить' if action.k1 > 0 else 'продать'} {abs(action.k1)} пак.")
    if action.k2:
        parts.append(f"ЦБ2: {'купить' if action.k2 > 0 else 'продать'} {abs(action.k2)} пак.")
    if action.k3:
        parts.append(f"Деп.: {'купить' if action.k3 > 0 else 'продать'} {abs(action.k3)} пак.")
    return ", ".join(parts) if parts else "без действий"


class PortfolioSolver:
    def __init__(self, data: Dict, free_cash: float) -> None:
        self.stages: List[List[Scenario]] = data["stages"]
        self.init = data["init"]
        self.commission = data["comm"]
        self.minimums = data["mins"]
        self.cash0 = free_cash

        # Package sizes = 25% of initial values
        self.pkg = {
            "cb1": self.init["cb1"] / 4,
            "cb2": self.init["cb2"] / 4,
            "dep": self.init["dep"] / 4,
        }

        self.min_p = {
            "cb1": int(ceil(self.minimums["cb1"] / self.pkg["cb1"])),
            "cb2": int(ceil(self.minimums["cb2"] / self.pkg["cb2"])),
            "dep": int(ceil(self.minimums["dep"] / self.pkg["dep"])),
        }

        # Cash costs/proceeds per package in UNITs
        self.buy_u = {
            k: int(round((self.pkg[k] * (100 + self.commission.get(k, 0)) / 100) / UNIT))
            for k in self.pkg
        }
        self.sell_u = {
            k: int(round((self.pkg[k] * (100 - self.commission.get(k, 0)) / 100) / UNIT))
            for k in self.pkg
        }

        # Expected multipliers for stage 3
        self.exp_m3 = {
            "cb1": sum(sc.prob * sc.m_cb1 for sc in self.stages[2]) / 100.0,
            "cb2": sum(sc.prob * sc.m_cb2 for sc in self.stages[2]) / 100.0,
            "dep": sum(sc.prob * sc.m_dep for sc in self.stages[2]) / 100.0,
        }

    def _apply_action_one(self, holding: int, cash: int, k: int, key: str) -> Tuple[int, int]:
        if k >= 0:
            return holding + k, cash - k * self.buy_u[key]
        return holding + k, cash + (-k) * self.sell_u[key]

    def feasible_actions(self, p1: int, p2: int, p3: int, cash: int) -> List[Action]:
        max_sell1 = max(0, p1 - self.min_p["cb1"])
        max_sell2 = max(0, p2 - self.min_p["cb2"])
        max_sell3 = max(0, p3 - self.min_p["dep"])

        max_cash = cash + max_sell1 * self.sell_u["cb1"] + max_sell2 * self.sell_u["cb2"] + max_sell3 * self.sell_u["dep"]
        max_buy1 = max_cash // self.buy_u["cb1"] if self.buy_u["cb1"] > 0 else 0
        max_buy2 = max_cash // self.buy_u["cb2"] if self.buy_u["cb2"] > 0 else 0
        max_buy3 = max_cash // self.buy_u["dep"] if self.buy_u["dep"] > 0 else 0

        actions: List[Action] = []
        for k1 in range(-max_sell1, max_buy1 + 1):
            p1n, cash1 = self._apply_action_one(p1, cash, k1, "cb1")
            if cash1 < 0 or p1n < self.min_p["cb1"]:
                continue
            for k2 in range(-max_sell2, max_buy2 + 1):
                p2n, cash2 = self._apply_action_one(p2, cash1, k2, "cb2")
                if cash2 < 0 or p2n < self.min_p["cb2"]:
                    continue
                for k3 in range(-max_sell3, max_buy3 + 1):
                    p3n, cash3 = self._apply_action_one(p3, cash2, k3, "dep")
                    if cash3 < 0 or p3n < self.min_p["dep"]:
                        continue
                    actions.append(Action(k1, k2, k3))
        return actions

    @lru_cache(maxsize=None)
    def solve_stage3(self, p1: int, p2: int, p3: int, cash: int) -> Tuple[int, Action]:
        # Greedy: sell assets with expected value < cash, then buy best asset by ratio
        action = Action(0, 0, 0)
        k1 = k2 = k3 = 0

        # Expected value per package (money)
        ev1 = self.pkg["cb1"] * self.exp_m3["cb1"] / 100.0
        ev2 = self.pkg["cb2"] * self.exp_m3["cb2"] / 100.0
        ev3 = self.pkg["dep"] * self.exp_m3["dep"] / 100.0

        # Sell if better to hold cash
        if self.sell_u["cb1"] * UNIT > ev1:
            sell = max(0, p1 - self.min_p["cb1"])
            p1 -= sell
            cash += sell * self.sell_u["cb1"]
            k1 -= sell
        if self.sell_u["cb2"] * UNIT > ev2:
            sell = max(0, p2 - self.min_p["cb2"])
            p2 -= sell
            cash += sell * self.sell_u["cb2"]
            k2 -= sell
        if self.sell_u["dep"] * UNIT > ev3:
            sell = max(0, p3 - self.min_p["dep"])
            p3 -= sell
            cash += sell * self.sell_u["dep"]
            k3 -= sell

        # Choose best asset to buy (ratio > 1)
        candidates = []
        if ev1 > self.buy_u["cb1"] * UNIT:
            candidates.append((ev1 / (self.buy_u["cb1"] * UNIT), "cb1"))
        if ev2 > self.buy_u["cb2"] * UNIT:
            candidates.append((ev2 / (self.buy_u["cb2"] * UNIT), "cb2"))
        if ev3 > self.buy_u["dep"] * UNIT:
            candidates.append((ev3 / (self.buy_u["dep"] * UNIT), "dep"))

        if candidates:
            _, best = max(candidates, key=lambda x: x[0])
            buy = cash // self.buy_u[best]
            if buy > 0:
                cash -= buy * self.buy_u[best]
                if best == "cb1":
                    p1 += buy
                    k1 += buy
                elif best == "cb2":
                    p2 += buy
                    k2 += buy
                else:
                    p3 += buy
                    k3 += buy

        action = Action(k1, k2, k3)
        total = p1 * self.pkg["cb1"] * self.exp_m3["cb1"] / 100.0
        total += p2 * self.pkg["cb2"] * self.exp_m3["cb2"] / 100.0
        total += p3 * self.pkg["dep"] * self.exp_m3["dep"] / 100.0
        total += cash * UNIT
        return int(round(total)), action

    @lru_cache(maxsize=None)
    def solve(self, stage: int, p1: int, p2: int, p3: int, cash: int) -> Tuple[int, Action]:
        if stage == 3:
            return self.solve_stage3(p1, p2, p3, cash)

        best_val = -10**18
        best_action = Action(0, 0, 0)

        for action in self.feasible_actions(p1, p2, p3, cash):
            p1a, cash1 = self._apply_action_one(p1, cash, action.k1, "cb1")
            p2a, cash2 = self._apply_action_one(p2, cash1, action.k2, "cb2")
            p3a, cash3 = self._apply_action_one(p3, cash2, action.k3, "dep")
            cash3 = round_cash_units(cash3)
            if cash3 < 0:
                continue

            exp_val_num = 0
            for sc in self.stages[stage - 1]:
                p1s = apply_multiplier_to_packages(p1a, self.pkg["cb1"], sc.m_cb1)
                p2s = apply_multiplier_to_packages(p2a, self.pkg["cb2"], sc.m_cb2)
                p3s = apply_multiplier_to_packages(p3a, self.pkg["dep"], sc.m_dep)
                val, _ = self.solve(stage + 1, p1s, p2s, p3s, cash3)
                exp_val_num += sc.prob * val

            exp_val = (exp_val_num + 50) // 100
            if exp_val > best_val:
                best_val = exp_val
                best_action = action

        return best_val, best_action

    def run(self) -> None:
        p1 = int(round(self.init["cb1"] / self.pkg["cb1"]))
        p2 = int(round(self.init["cb2"] / self.pkg["cb2"]))
        p3 = int(round(self.init["dep"] / self.pkg["dep"]))
        cash = int(round(self.cash0 / UNIT))
        cash = round_cash_units(cash)

        total, act1 = self.solve(1, p1, p2, p3, cash)
        print("Expected final wealth (Bayes):", total)
        print("Stage 1 action:", action_to_text(act1))

        p1a, cash1 = self._apply_action_one(p1, cash, act1.k1, "cb1")
        p2a, cash2 = self._apply_action_one(p2, cash1, act1.k2, "cb2")
        p3a, cash3 = self._apply_action_one(p3, cash2, act1.k3, "dep")
        cash3 = round_cash_units(cash3)

        for i, sc in enumerate(self.stages[0], start=1):
            p1s = apply_multiplier_to_packages(p1a, self.pkg["cb1"], sc.m_cb1)
            p2s = apply_multiplier_to_packages(p2a, self.pkg["cb2"], sc.m_cb2)
            p3s = apply_multiplier_to_packages(p3a, self.pkg["dep"], sc.m_dep)
            _, act2 = self.solve(2, p1s, p2s, p3s, cash3)
            print(f"Stage 2 action (scenario {i}):", action_to_text(act2))

            p1b, cash4 = self._apply_action_one(p1s, cash3, act2.k1, "cb1")
            p2b, cash5 = self._apply_action_one(p2s, cash4, act2.k2, "cb2")
            p3b, cash6 = self._apply_action_one(p3s, cash5, act2.k3, "dep")
            cash6 = round_cash_units(cash6)

            for j in range(1, 4):
                _, act3 = self.solve(3, p1b, p2b, p3b, cash6)
                print(f"  Stage 3 action (scenario {i}.{j}):", action_to_text(act3))


if __name__ == "__main__":
    EXCEL_PATH = Path(__file__).with_name("Данные для постановки задачи.xlsx")
    FREE_CASH = 600.0  # если нужно, поменяйте

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Не найден файл: {EXCEL_PATH}")

    data = read_excel(str(EXCEL_PATH))
    solver = PortfolioSolver(data, FREE_CASH)
    solver.run()
